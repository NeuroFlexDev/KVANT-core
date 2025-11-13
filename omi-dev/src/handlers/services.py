import requests
import json
import openpyxl
import os
import re
import shutil
from urllib.parse import urljoin
from fastapi import HTTPException, UploadFile
from starlette.responses import Response

from settings import SERVER_OMIAPI
from handlers import BaseHandler
from services.excel import ExportExcel
from services.word import ExportWord
from services.pdf import omi
from models.requests.export import ReportFormat
from models.requests import UpdateEstimate, CalcType
from dc_core.models.db.user import DBUser
from handlers._calc import create_estimate

__all__ = ("ServicesHandler",)


class ServicesHandler(BaseHandler):
    # def update_estimate(self, data: UpdateEstimate, current_user: DBUser):
    #     rs = self.db_layer.registries.get_questionnaire_object_by_id(id=data.id, user_id=current_user.id)
    #     if not rs:
    #         return
    #
    #     estimate = rs.get("estimate")
    #     if estimate["result"]["code"] != 0:
    #         return
    #
    #     prices = data.prices
    #
    #     sections = estimate["sections"]
    #     total_cost = 0
    #     for s, section in enumerate(sections):
    #         pos_1 = 0
    #         section_cost = 0
    #         element_types = section.get("element_types")
    #         for e, element_type in enumerate(element_types):
    #             pos_1 += 1
    #             pos_2 = 0
    #             predicts_new = []
    #             element_cost = 0
    #             predicts = element_type.get("predicts")
    #             for p, predict in enumerate(predicts):
    #                 pos_2 += 1
    #                 element_price = predict.get("price")
    #                 for price in prices:
    #                     price = dict(price)
    #                     if price.get("num") == predict.get("num"):
    #                         element_price = price.get("value")
    #                         estimate["sections"][s]["element_types"][e]["predicts"][p]["price"] = element_price
    #                 element_cost += element_price
    #
    #             estimate["sections"][s]["element_types"][e]["element_cost"] = element_cost
    #             section_cost += element_cost
    #
    #         total_cost += section_cost
    #         estimate["sections"][s]["section_cost"] = section_cost
    #
    #     estimate["result"]["total_cost"] = total_cost
    #
    #     rs = self.db_layer.registries.recording_response(
    #         id=data.id,
    #         response=None,
    #         estimate=json.dumps(estimate)
    #     )
    #
    #     if not rs:
    #         raise HTTPException(404, f"Строка с ID={data.id} не найдена")
    #
    #     try:
    #         self.db_layer.commit()
    #     except Exception as e:
    #         self.db_layer.rollback()
    #         raise e
    #
    #     return estimate

    def get_request(self, id: int, current_user: DBUser):
        if 4 in current_user.role_ids:
            user_id = None
        else:
            user_id = current_user.id

        user = self.db_layer.users_omi.get_user_by_email(email=current_user.email)
        calc_number = user.get("calc_number") or 0
        if calc_number < 1:
            raise HTTPException(402, f"Небходима оплата сервиса")


        data = self.db_layer.registries.get_questionnaire_object_by_id(id=id, calc_type=None, user_id=user_id)
        if not data:
            raise HTTPException(404, f"Нет данных")
        params = []
        for js in data.get("params"):
            el = dict()
            for key, val in js.items():
                if key == "desc":
                    el["key"] = js[key].rstrip(".")
                else:
                    el["value"] = js[key]
            params.append(el)

        # Запрос к сервису
        sess = requests.Session()
        url = urljoin(SERVER_OMIAPI, "request/")
        request_body = {
            "id": data.get("id"),
            "params": params,
        }

        # return request_body

        rss = requests.post(
            url,
            json=request_body,
            headers={"Content-Type": "application/json"},
        )

        if rss.status_code == 200:
            response = rss.json()
            estimate = create_estimate(self, id, response, False, current_user)
        else:
            return rss.content

        r = {
            "id": data.get("id"),
            "params": response
        }

        self.db_layer.users_omi.replenish_account(
            email=current_user.email,
            value=-1,
            user=current_user.email
        )

        self.db_layer.logs.create_log(
            table_name="dir.questionnaire_object",
            record_id=id,
            action="calculation",
            user_id=current_user.id
        )

        try:
            self.db_layer.commit()
        except Exception as e:
            self.db_layer.rollback()
            raise e

        return r

    def get_response(self, id: int, is_filter_null_values: bool, current_user: DBUser):
        if 4 in current_user.role_ids:
            user_id = None
        else:
            user_id = current_user.id

        data = self.db_layer.registries.get_questionnaire_object_by_id(id=id, calc_type=None, user_id=user_id)
        if not data:
            return

        response = data.get("response")
        if response == None or response["result"]["code"] != 0:
            return

        estimate = data.get("estimate")
        # if estimate == None:
        estimate = create_estimate(self, id, response, False, current_user)

        return estimate

    def export(self, id: int, fmt: ReportFormat, calc_type: CalcType, current_user: DBUser):
        if 4 in current_user.role_ids:
            user_id = None
        else:
            user_id = current_user.id

        data = self.get_response(id, False, current_user)

        rs = self.db_layer.registries.get_questionnaire_object_by_id(id=id, calc_type=calc_type, user_id=user_id)
        body = rs.get("params")
        params = {}
        for b in body:
            params.update(b)
        params.pop("desc")

        if fmt == ReportFormat.EXCEL:
            body = ExportExcel().omi(id, data, calc_type)
            return Response(
                body,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=questionnaire_{id}.xlsx"},
            )
        elif fmt == ReportFormat.WORD:
            body = ExportWord().omi(id, data, params, calc_type)
            return Response(
                body,
                media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                headers={"Content-Disposition": f"attachment; filename=questionnaire_{id}.docx"},
            )
        elif fmt == ReportFormat.PDF:
            body = omi(id, data, params, calc_type)

            return Response(
                body,
                media_type="application/pdf",
                headers={"Content-Disposition": f"attachment; filename=questionnaire_{id}.pdf"},
            )
        else:
            return


    def import_file(self, file: UploadFile, object_numbers: str, questionnaire_id: int, update: bool, current_user: DBUser):
        upload_dir = os.path.join(os.getcwd(), "tmp", "uploads")
        # print(upload_dir)

        if not os.path.exists(upload_dir):
            os.makedirs(upload_dir)

        dest = os.path.join(upload_dir, file.filename)
        # print(dest)

        # copy the file contents
        with open(dest, "wb") as buffer:
            shutil.copyfileobj(file.file, buffer)

        try:
            wookbook = openpyxl.load_workbook(dest, data_only=True)
        except Exception as e:
            raise HTTPException(404, str(e))
        worksheet = wookbook.active

        pos = []
        for col in worksheet.iter_cols(1, 1):
            for i in range(0, worksheet.max_row):
                pos.append(col[i].value)
        names = []
        for col in worksheet.iter_cols(2, 2):
            for i in range(0, worksheet.max_row):
                names.append(col[i].value)

        if object_numbers:
            object_numbers_array = [val.strip() for val in object_numbers.split(",")]
        else:
            object_numbers_array = []

        log_full = {}

        for col in worksheet.iter_cols(10, worksheet.max_column):
            # Наименование и адрес объекта
            object_num = None
            object_name = None
            object_address = None
            for i in range(0, 10):
                if pos[i] == "№ п.п.":
                    object_num = str(col[i].value)
                elif pos[i] == "1.1":
                    object_name = str(col[i].value)
                    if object_name and object_name.strip() == "Жилой дом" and object_num:
                        object_name = f"{object_num} - {object_name}"
                elif pos[i] == "1.2":
                    object_address = str(col[i].value)
                if object_name and object_address:
                    break

            if object_num == "None":
                return log_full

            if object_numbers and object_num not in object_numbers_array:
                continue

            print("==================", object_num, "==================")

            if object_name == None or object_address == None:
                raise HTTPException(400, "В файле не заполнена информация об объектах")
            object_rs = self.db_layer.registries.get_object_by_params(
                name=object_name,
                address=object_address,
                user_id=current_user.id
            )
            if not object_rs:
                object_rs = self.db_layer.registries.create_object(
                    name=object_name,
                    address=object_address,
                    user_id=current_user.id
                )

                try:
                    self.db_layer.commit()
                except Exception as e:
                    self.db_layer.rollback()
                    raise e
            # print(">>>", object_rs)
            object_id = object_rs.get("id")

            qo = self.db_layer.registries.get_questionnaire_object_by_params(
                name=object_num,
                object_id=object_id,
                questionnaire_id=questionnaire_id,
                user_id=current_user.id,
            )
            # print(qo)
            if not qo:
                qo = self.db_layer.registries.create_questionnaire_object(
                    name=object_num,
                    questionnaire_id=questionnaire_id,
                    object_id=object_id,
                    user_id=current_user.id,
                )
                qo_new = True
            else:
                if not update:
                    print("passed")
                    log_full[object_num] = "Пропущен"
                    continue
                qo_new = False

                try:
                    self.db_layer.commit()
                except Exception as e:
                    self.db_layer.rollback()
                    raise e
            qo_id = qo.get("id")

            rs = self.db_layer.registries.get_questions_by_questionnaire_id(questionnaire_id=questionnaire_id)
            # print(rs)





            params = []
            global log
            log = {}

            a4 = None
            a5 = None
            a18 = []

            for i in range(0, worksheet.max_row):
                value = col[i].value
                if "Зд." in str(value):
                    value = re.compile(r"\s+Зд.").sub("\nЗд.", value).strip()

                    # Проверка номера здания
                    results = re.findall("Зд.\d+", value)
                    for result in results:
                        n = int(result.replace("Зд.", ""))
                        # print("++++++++++++++", n, a5)
                        if int(result.replace("Зд.", "")) > a5:
                            log[pos[i]] = f"Недопустимое значение номера здания: {n}"
                            value = None
                            break
                    if value == None:
                        continue

                if "секц" in str(value) and a18:
                    # Проверка номера секции
                    tmp = value.replace(" ", "")
                    regex = "(Зд)[.]*\s*(\d+)[\s_-]*(\d*)\s*(секц)[.]*(\d*)"
                    subst = "\\1.\\2_\\4.\\3\\5"
                    tmp = re.sub(regex, subst, tmp)
                    # print(tmp)

                    results = re.findall("Зд.(\d+)_секц.(\d+)", tmp)
                    for result in results:
                        building = int(result[0])
                        sections = int(result[1])
                        # print("%%%%%%%%%%%%%%%%", building, sections, a18)
                        if sections > a18[building - 1]:
                            log[pos[i]] = f"Недопустимое значение номера секции: {sections}"
                            value = None
                            break
                    if value == None:
                        continue


                p = pos[i] if not pos[i] else pos[i] + "."
                param_a = None
                for r in rs:
                    if r.get("pos") == p:
                        param_a = r.get("param")
                        preset_values = r.get("values")
                        datatype = r.get("datatype")
                        print(param_a, pos[i], value, datatype)
                        if param_a == "a4":
                            a4 = value
                            # print("--a4 =", a4)
                        elif param_a == "a5":
                            a5 = value if value > 1 else 1
                            # print("--a5 =", a5)
                            if a4:
                                if a4.lower() == "нет" and a5 > 1:
                                    log[pos[i]] = f"Значение '{value}' не допустимо при текущем значении параметра 2.1.2"
                                    value = None
                                    break
                                elif a4.lower() == "да" and a5 < 1:
                                    log[pos[i]] = f"Значение '{value}' не допустимо при текущем значении параметра 2.1.2"
                                    value = None
                                    break

                        if type(preset_values) is list:
                            if datatype in ["list_s", "list_m"]:
                                # игнорируем нули в ячейках, где должен быть список
                                if value == 0:
                                    value = None
                                    break
                                vals = [v for v in value.split("\n")]
                                # print(">>>", vals)
                                if "Зд." in str(vals):
                                    log[pos[i]] = f"Значение '{value}' не допустимо. Необходимо указать одно значение"
                                    value = None
                                    break
                                value = []
                                for val in vals:
                                    found = False
                                    for preset_value in preset_values:
                                        if val.replace('"', '').replace(" ", "").strip("- .;«»").lower() in preset_value.replace(" ", "").lower():
                                            found = True
                                            if datatype == "list_s":
                                                value = preset_value
                                            else:
                                                value.append(preset_value)
                                            break
                                    if not found:
                                        log[pos[i]] = f"Значение '{val}' не допустимо"
                                        # print(f"===================ОЛ #{object_num}:", pos[i], val)

                                if param_a in ["a32", "a33", "a34", "a35"]:
                                    new_val = []
                                    for s in a18:
                                        sect = []
                                        for n in range(s):
                                            sect.append(value)
                                        new_val.append(sect)
                                    value = new_val
                                    # print("********", value)
                                # if datatype == "list_s":
                                #     value = None
                                # else:
                                #     value = []
                                # for preset_value in preset_values:
                                #     if preset_value.lower() in vals:
                                #         # print(pos[i], "---", preset_value)
                                #         if datatype == "list_s":
                                #             value = preset_value
                                #         else:
                                #             value.append(preset_value)
                            elif datatype == "dict_m":
                                if value == 0:
                                    value = None
                                    break
                                vals = value.lower()
                                value = []
                                for preset_value in preset_values:
                                    for key, val in preset_value.items():
                                        if key in vals:
                                            value.append(preset_value)
                                # print(pos[i], ">>>>>>>", value)
                        else:
                            if datatype == "str":
                                if type(value) is float:
                                    value = str(value).replace(".", ",")
                                elif type(value) is int:
                                    value = str(value)
                            elif datatype == "int":
                                # print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
                                if type(value) is float:
                                    value = round(value)
                                # elif param_a == "a45":
                                #     value = [value]
                            elif datatype == "float":
                                if type(value) is float:
                                    value = round(value * 10000) / 10000
                            else:
                                # print(">>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>")
                                value = None

                            if param_a in ["a18", "a19", "a20", "a21", "a31", "a44", "a73"]:
                                if param_a != "a73":
                                    regex = "\s*-\s*"
                                    subst = " - "
                                    value = re.sub(regex, subst, value)
                                if param_a == "a18":
                                    value = value.replace(" ", "")
                                    regex = "(Зд)[.]*\s*(\d+)[\s_-]*(\d*)\s*(секц)[.]*(\d*)"
                                    subst = "\\1.\\2 - \\4.\\5\\3"
                                    value = re.sub(regex, subst, value)
                                else:
                                    regex = "(Зд)[.]*\s*(\d+)[\s_-]*(\d*)\s*(секц)[.]*(\d*)"
                                    subst = "\\1.\\2_\\4.\\3\\5"
                                    value = re.sub(regex, subst, value)

                                if "Зд." in str(value):
                                    # print("%%%%%%", value)
                                    value = self.parsing(value, param_a, pos[i])
                                elif str(value) == "0":
                                    value = None
                                else:
                                    log[pos[i]] = f"Значение '{value}' не допустимо. Необходимо указать в разрезе Здания/Секции"
                                    value = None

                                if param_a == "a18":
                                    a18 = value
                                    # print("=====!!!=====", a18)
                                    if type(a18) is not list:
                                        log[pos[i]] = f"Неверное значение параметра '{value}'"

                            elif param_a == "a45":
                                if value == 0:
                                    value = None
                                elif str(value).isdigit():
                                    value = [value]
                                else:
                                    log[pos[i]] = f"Значение '{value}' не допустимо. Необходимо указать целое значение"
                                    value = None
                            # elif param_a == "a73":
                            #     pass
                                # log[pos[i]] = f"Значение '{value}' не допустимо. Необходимо указать в разрезе Здания/Секции"
                                # value = None
                            elif param_a == "a78":
                                # print(value)
                                value = self.parsing_a78(value)
                            elif "Зд." in str(value):
                                value = self.parsing(value, param_a, pos[i])

                        break
                if param_a == None:
                    continue
                    # raise HTTPException(404, f"Параметр {pos[i]} не найден")

                print("$", value)
                params.append(
                    {
                        param_a: value,
                        "desc": pos[i]
                    }
                )

            if log:
                log_full[object_num] = log
            else:
                log_full[object_num] = "Успешно"

            if qo_new or update:
                rs = self.db_layer.registries.update_questionnaire_object(
                    id=qo_id,
                    state_id=0,
                    name=object_num,
                    questionnaire_id=questionnaire_id,
                    object_id=object_id,
                    params=json.dumps(params),
                    response=None,
                    estimate=None,
                    user_id=current_user.id,
                )

                try:
                    self.db_layer.commit()
                except Exception as e:
                    self.db_layer.rollback()
                    raise e

        return log_full


    def parsing_a78(self, value):
        # print("<<<<<<<<<<<<<<<<<", value)
        if type(value) is int:
            return [value]
        if "," in value:
            value = [int(val) for val in value.split(",")]
        elif str(value).lstrip("-").isdigit():
            return [value]
        elif "-" in value:
            value = [int(val) for val in value.split("-")]
        else:
            return [value]
        return [i for i in range(value[0], value[1] + 1)]

    def parsing(self, value, param, pos):
        value_orig = value
        # print("==================")
        if param == "a44":
            pass
        elif param == "a73":
            value = value.replace("- ", "_")
            # print(">>", value)
        else:
            for val in value.split("\n"):
                if " - " not in val:
                    log[pos] = f"Значение '{val}' не допустимо. Отсутствует разделитель ' - '"
                    return None


        pattern = "[А-Яа-я.]"
        value = re.sub(pattern, '', value)
        # print(">>>", value)
        value = [
            [
                [x.strip() if not x.strip().isdigit() else int(x.strip()) for x in vv.split("_")]

                for vv in v.split(" - ")
            ]
            for v in value.split("\n")

        ]
        # print("VALUE:", value)

        if param == "a44":
            # print(value)
            sections = "Здание %s секция %s"
            try:
                return [sections % tuple(val[0]) for val in value]
            except:
                log[pos] = f"Не получилось разобрать строку '{value_orig}'"
                return None

        building = []
        section = []
        building_num = 0
        section_num = 0

        for val in value:
            # print("---", len(val[0]), val[0])
            if len(val[0]) == 1:
                # print(val, type(val), val[1][0])
                v = val[1][0]
                if type(v) is str:
                    if "," in v:
                        v = (v.replace(",", "."))
                    else:
                        v = int(v)
                building.insert(building_num, v)
                building_num += 1

            elif len(val[0]) == 2:
                # print(building_num, section_num)

                if building_num != val[0][0]:
                    if section:
                        building.insert(building_num, section)
                    building_num += 1
                    section_num = 0
                    section = []

                # print(val, type(val))
                v = val[1][0]
                if not v:
                    log[pos] = f"Не получилось разобрать строку '{value_orig}'"
                    return None
                # print("***", val, v)
                if type(v) is str:
                    if "," in v:
                        v = (v.replace(",", "."))
                    else:
                        v = int(v)
                elif param == "a73":
                    v = [v]
                section.insert(section_num, v)
                section_num += 1
            else:
                if building_num != val[0][0]:
                    if section:
                        building.insert(building_num, section)
                    building_num += 1
                    section_num = 0
                    section = []

                # print(val, type(val), val[0])
                v = val[0][2]
                # print("***", v)
                v = self.parsing_a78(v)
                # print("***", v)

                # section.insert(section_num, [int(x) for x in val[0][2].split(",")])
                section.insert(section_num, v)
                section_num += 1
            # print("B:", building, "S:", section)

        if len(val[0]) > 1:
            building.insert(building_num, section)

        # print("===", building)

        return building
